"""
biscotti.export
~~~~~~~~~~~~~~~
Export bulk run results to CSV, TSV, or XLSX format.
"""
from __future__ import annotations

import csv
import io
from typing import Sequence

from .models import RunLog


# Short scannable metrics come first, then long text columns (Output, Judge
# Reasoning) last — keeps the first screen of the spreadsheet readable.
_HEADERS_BASE = [
    "Test Case", "Model", "Temperature", "Reasoning Effort", "Outcome",
    "Latency (ms)", "Input Tokens", "Output Tokens", "Estimated Cost",
    "Output", "Error Message",
]

_HEADERS_WITH_SCORE = _HEADERS_BASE + ["Score", "Judge Reasoning"]


def _run_to_row(run: RunLog, include_score: bool) -> list:
    outcome_val = run.outcome.value if hasattr(run.outcome, "value") else str(run.outcome)
    row = [
        run.test_case_name or "",
        run.model_used,
        run.temperature if run.temperature is not None else "",
        run.reasoning_effort or "",
        outcome_val,
        run.latency_ms,
        run.input_tokens,
        run.output_tokens,
        f"{run.estimated_cost:.6f}" if run.estimated_cost is not None else "",
        run.output or "",
        run.error_message or "",
    ]
    if include_score:
        row.append(run.score if run.score is not None else "")
        row.append(run.score_reasoning or "")
    return row


def generate_export(runs: Sequence[RunLog], format: str = "csv", include_score: bool = False) -> bytes:
    headers = _HEADERS_WITH_SCORE if include_score else _HEADERS_BASE
    rows = [_run_to_row(r, include_score) for r in runs]
    if format == "tsv":
        return _generate_delimited(headers, rows, delimiter="\t")
    elif format == "xlsx":
        return _generate_xlsx(headers, rows)
    else:
        return _generate_delimited(headers, rows, delimiter=",")


def _generate_delimited(headers: list[str], rows: list[list], delimiter: str) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _generate_xlsx(headers: list[str], rows: list[list]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Run Results"
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
